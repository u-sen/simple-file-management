# coding:utf8

import datetime
import os
# from tkinter import ttk
from tkinter import *
from tkinter import filedialog
from tkinter import messagebox

import openpyxl

# from typing import List, Any

# import xlwings

Language = 'zh'
Fdpath = ""
Fd = 'folder'
Workpath = os.getcwd()


# print(next(os.walk(Workpath))[2])

class app:
    global Fdpath
    # str变量
    s_label_filename = "读取文件名"
    s_label_Fdname = "目录:"
    ListBox_var = ""  # 选择的文件名
    # list 变量
    l_entry = [x for x in 12 * " "]  # entry列表
    # num 变量
    n_sign = 0  # 功能选择
    n_activate = -1  # listbox激活号
    # 布尔量
    b_event_activate = False
    b_top_living = False
    b_init_Fd = False

    def __init__(self):
        self.pr = Fd_editor()
        # self.log = log_app()
        self.init_wd_name = Tk()
        self.set_init_window()

    # 运行窗体
    def run(self):
        global Fdpath
        """运行窗体"""
        # self.click_button_Refresh()
        if E_file.cfg_ini("Fd_name_cfg", "r")[0]: Fdpath = E_file.cfg_ini("Fd_name_cfg", "r")[1]
        self.init_wd_name.mainloop()

    # 选择列表事件
    def Event_listbox_var(self, event):
        self.n_activate = self.listbox_fd.curselection()
        if self.n_activate != ():
            self.ListBox_var = self.listbox_fd.get(self.n_activate)
            self.listbox_fd.activate(self.n_activate)
            self.reinput()
            list_temp = self.showinxlsx()
            if list_temp == self.l_entry:
                self.n_activate = -1
                return
            self.show_entry(self.showinfd() + list_temp[7:12])
            self.show_text(list_temp[6])
            self.b_event_activate = True
        self.n_activate = -1
        # print(self.ListBox_var)

    # 获取当前日期快捷建 Ctrl + ; 事件
    def Event_date_new(self, event):
        if event.widget == self.entry_date:
            self.entry_date.delete(0, "end")
            self.entry_date.insert(0, self.auto_date())
        if event.widget == self.text_content:
            self.text_content.insert("insert", self.auto_date())

    def Even_window(self, event):
        if not self.b_init_Fd:
            self.click_button_Refresh()
            self.b_init_Fd = True

    # 获取当前日期
    @staticmethod
    def auto_date():
        datel = datetime.datetime.now()
        return datel.strftime('%Y/%m/%d')
        # eval("self.{}.delete".format(widget))(0, "end")
        # eval("self.{}.insert".format(widget))("insert", datel.strftime('%Y/%m/%d'))

    # 格式显示文件夹名称到输入窗口
    def showinfd(self):
        a, b, c, d = self.ListBox_var[:self.ListBox_var.find("_")], \
                     self.ListBox_var[self.ListBox_var.find("_") + 1:self.ListBox_var.find(" ")], \
                     self.ListBox_var[self.ListBox_var.find(" ") + 1:self.ListBox_var.rfind(" ")], \
                     self.ListBox_var[self.ListBox_var.rfind(" ") + 1:]
        a = self.formatting_id(a)
        b = self.formatting_serialnumber(b)
        d = self.formatting_date(d, inFd=True)
        return [a, b, c, d]

    # 格式显示表格内容到输入窗口（所选编号）
    def showinxlsx(self):
        check_temp = self.pr.set_Xlsx("open")
        id_temp = int(self.ListBox_var[:self.ListBox_var.find("_")])
        if check_temp[0]:
            check_list = self.pr.read_el(id_temp)
            if check_list == self.l_entry:
                messagebox.showerror("表格读取错误", "不存在编号(" + str(id_temp) + ") 的行\n>>请检查表格")
            return check_list
        else:
            messagebox.showerror("表格读取错误", "请检查表格名：\n" + check_temp[1])
            return [[x for x in 12 * " "]]

    # 格式化编号ID
    @staticmethod
    def formatting_id(idin):
        if len(idin) == 1:
            return "00" + idin
        elif len(idin) == 2:
            return "0" + idin
        elif len(idin) >= 3:
            return idin
        else:
            return "0000"

    # 格式序列号
    def formatting_serialnumber(self, serialnumberin):
        if serialnumberin != "":
            return serialnumberin.upper()
        else:
            messagebox.showerror("内容错误", "没有输入序列号")
            return ""

    # 格式化日期
    def formatting_date(self, datein, inFd=False, newFd=False):
        try:
            if inFd:
                if datein.find("-") == -1:
                    datel = datetime.datetime.strptime(datein, '%Y%m%d')
                else:
                    datel = datetime.datetime.strptime(datein, '%Y-%m-%d')
            else:
                datel = datetime.datetime.strptime(datein, '%Y/%m/%d')
            if newFd:
                return datel.strftime('%Y-%m-%d')
            else:
                return datel.strftime('%Y/%m/%d')
        except ValueError:
            # self.E_file.w_log()
            messagebox.showerror("日期错误", "请输入正确的格式：年/月/日")
            return ""

    # 设置窗口
    def set_init_window(self):
        # 主窗口属性
        self.init_wd_name.title("问题管理表")
        self.init_wd_name.geometry("510x550+500+100")
        self.init_wd_name.resizable(0, 0)
        self.init_wd_name.bind('<FocusIn>', self.Even_window)
        self.init_wd_name.iconbitmap(Fdpath + "Edititem.ico")
        # 元素标签
        self.label_fd = Label(self.init_wd_name, text=self.s_label_Fdname, anchor="w", width=60)
        self.label_fd.grid(row=0, column=0, columnspan=3, sticky="w", pady='5px')
        self.label_filename = Label(self.init_wd_name, text=self.s_label_filename, anchor="w", width=60)
        self.label_filename.grid(row=3, column=0, sticky="w", columnspan=3, pady='5px')
        self.label_Id = Label(self.init_wd_name, text="编号")
        self.label_Id.grid(row=4, column=0, sticky="w")
        self.label_serialnumber = Label(self.init_wd_name, text="序列号/工位号").grid(row=4, column=1, sticky="w")
        self.label_problem = Label(self.init_wd_name, text="问题名").grid(row=4, column=2, sticky="w")
        self.label_date = Label(self.init_wd_name, text="日期").grid(row=6, column=0, sticky="w", pady='5px')
        self.label_raiseAquestion = Label(self.init_wd_name, text="提出人").grid(row=6, column=1, sticky="w")
        self.label_ProblemHandling = Label(self.init_wd_name, text="处理人").grid(row=6, column=2, sticky="w")
        self.label_state = Label(self.init_wd_name, text="状态").grid(row=8, column=0, sticky="w", pady='5px')
        self.label_frequency = Label(self.init_wd_name, text="频次").grid(row=8, column=1, sticky="w")
        self.label_remark = Label(self.init_wd_name, text="备注").grid(row=8, column=2, sticky="w")
        self.label_content = Label(self.init_wd_name, text="内容", width=60).grid(row=10, column=0, columnspan=3,
                                                                                sticky="w")
        # 元素列表
        # self.ListBox_var = StringVar()
        self.listbox_fd = Listbox(self.init_wd_name, height=5, width=60,
                                  exportselection=False)  # , listvariable=self.ListBox_var)
        self.listbox_fd.grid(row=1, column=0, rowspan=2, columnspan=3)
        self.listbox_fd.bind("<ButtonRelease-1>", self.Event_listbox_var)
        self.listbox_fd.bind("<Return>", self.Event_listbox_var)
        # cbox 元素
        # self.cbox = ttk.Combobox(self.init_wd_name)
        # self.cbox.grid(row=3,column=2)
        # Entry输入框元素
        self.list_entry = ["entry_id", "entry_serialnumber", "entry_problem", "entry_date", "entry_raiseAquestion",
                           "entry_ProblemHandling", "entry_state", "entry_frequency", "entry_remark"]
        self.entry_id = Entry(self.init_wd_name, width=4, state="disabled", )
        self.entry_id.grid(row=5, column=0, sticky="w", pady='5px')
        # self.entry_id.bind("<Return>", self.x)
        self.entry_serialnumber = Entry(self.init_wd_name, width=15)
        self.entry_serialnumber.grid(row=5, column=1, sticky="w")
        self.entry_problem = Entry(self.init_wd_name, width=30)
        self.entry_problem.grid(row=5, column=2, sticky="w")
        self.entry_date = Entry(self.init_wd_name, width=10)
        self.entry_date.grid(row=7, column=0, sticky="w", pady='5px')
        self.entry_raiseAquestion = Entry(self.init_wd_name, width=16)
        self.entry_raiseAquestion.grid(row=7, column=1, sticky="w")
        self.entry_ProblemHandling = Entry(self.init_wd_name, width=16)
        self.entry_ProblemHandling.grid(row=7, column=2, sticky="w")
        self.entry_state = Entry(self.init_wd_name, width=8)
        self.entry_state.grid(row=9, column=0, sticky="w", pady='5px')
        self.entry_frequency = Entry(self.init_wd_name, width=5)
        self.entry_frequency.grid(row=9, column=1, sticky="w")
        self.entry_remark = Entry(self.init_wd_name, width=30)
        self.entry_remark.grid(row=9, column=2, sticky="w")
        # 文本框元素
        self.text_content = Text(self.init_wd_name, width=60, height=10, wrap="word", autoseparators=False, undo=False)
        self.text_content.grid(row=11, column=0, columnspan=3)
        self.text_content.bind("<Control-KeyPress-;>", self.Event_date_new)
        # self.text_content.bind("<-->", self.Event_date_new)
        # 元素按键
        self.button_select = Button(self.init_wd_name, text="选择管理目录", width=10, command=self.click_button_select).grid(
            row=0, column=4)
        self.button_Refresh = Button(self.init_wd_name, text="刷新目录", width=10, command=self.click_button_Refresh).grid(
            row=1, column=4)
        self.button_openFd = Button(self.init_wd_name, text="打开文件夹", width=10, command=self.click_button_openFd).grid(
            row=2, column=4)
        self.button_open = Button(self.init_wd_name, text="打开表格", width=10, command=self.click_button_open).grid(row=3,
                                                                                                                 column=4)
        self.button_newFd = Button(self.init_wd_name, text="新建文件夹", width=10, command=self.click_button_newFd).grid(
            row=5, column=4)
        self.button_reFd = Button(self.init_wd_name, text="修改文件名", width=10, command=self.click_button_reFd).grid(row=4,
                                                                                                                  column=4)
        self.button_Ok = Button(self.init_wd_name, text="确认", width=10, command=self.click_button_Ok, bd=2)
        self.button_Cancel = Button(self.init_wd_name, text="取消", width=10, command=self.click_button_Cancel, bd=2)
        self.button_editor = Button(self.init_wd_name, text="编辑", width=10, command=self.click_button_editor).grid(
            row=10, column=4)
        self.button_formatFd = Button(self.init_wd_name, text="格式文件夹名", width=10, command=self.click_formatFd, bd=2)

    # 增加选择表格文件弹窗
    # 设置表格名
    def set_label_filename(self):
        liet_temp = self.pr.get_file_list()
        if E_file.cfg_ini("Xlsx_name_cfg", "r", Fdnamein=Fdpath)[0]:
            self.s_label_filename = E_file.cfg_ini("Xlsx_name_cfg", "r", Fdnamein=Fdpath)[1]
        elif len(liet_temp) > 1:
            E_file.w_log("app", "选择文件窗口>>打开")
            self.run_top()
        else:
            self.s_label_filename = self.pr.get_file_list()[0]
        self.pr.s_xlsx_name = self.s_label_filename
        self.label_filename.config(text=self.s_label_filename)

    # 设Ok/Cancel 按键位置
    def set_button_OKandCancel(self):
        self.button_Ok.grid(row=7, column=4)
        self.button_Cancel.grid(row=8, column=4)

    # Ok按键功能
    def click_button_Ok(self):
        E_file.w_log("app", "确认夹按键>>点击确认")
        b_Ok = True
        index_l = 0
        match self.n_sign:
            case 1:
                b_Ok = self.click_button_reFd(from_ok=True)
                index_l = self.n_activate
            case 2:
                b_Ok = self.click_button_newFd(from_ok=True)
            case 3:
                b_Ok = self.click_button_editor(from_ok=True)
            case 0:
                pass
        if b_Ok:
            self.click_button_Refresh()
            if self.n_sign == 1: self.listbox_fd.see(index_l)
            self.n_sign = 0
            self.button_Ok.grid_forget()
            self.button_Cancel.grid_forget()

        # self.init_wd_name.update()

    # Cancel按键功能
    def click_button_Cancel(self):
        E_file.w_log("app", "取消按键>>点击取消")
        self.n_sign = 0
        self.button_Ok.grid_forget()
        self.button_Cancel.grid_forget()
        self.button_formatFd.grid_forget()

    # 格式化文件夹名 更新表格链接
    def click_formatFd(self):
        E_file.w_log("app", "格式文件夹名>>点击格式文件夹名")
        if not messagebox.askokcancel("格式文件夹名", "将会格式化当前文件目录下的文件夹，及更新表格链接"): return
        list_temp = self.pr.Listinfolder()
        for i in list_temp:
            a, b, c, d = i[:i.find("_")], i[i.find("_") + 1:i.find(" ")], i[i.find(" "):i.rfind(" ")], i[i.rfind(
                " ") + 1:]
            a = self.formatting_id(a)
            b = self.formatting_serialnumber(b)
            while c.find(" ") != -1 and c.find(" ") == 0: c = c[c.find(" ") + 1:]
            d = self.formatting_date(d, inFd=True, newFd=True)
            newName = a + "_" + b + " " + c + " " + d
            self.pr.reFd(i, newName)
            val_temp = [int(a), newName]
            check_temp = self.pr.set_Xlsx("open")
            check_temp1 = self.pr.write_el(val_temp)
            if check_temp[0] and check_temp1[0]:
                check_temp3 = self.pr.set_Xlsx("save")
                self.click_button_Refresh()
                if not check_temp3[0]:
                    messagebox.showerror("表格保存错误", "请检查：\n" + check_temp3[1])
            elif not check_temp[0]:
                messagebox.showerror("表格打开错误", "请检查：\n" + check_temp[1])
                return False
            elif not check_temp1[0]:
                messagebox.showerror("表格写入错误", "请检查：\n" + check_temp1[1])
                return False

    # 选择工作路径
    def click_button_select(self):
        E_file.w_log("app", "选择路径按键>>点击选择路径")
        global Fdpath
        Fdpath = filedialog.askdirectory()
        if Fdpath != '':
            self.label_fd.config(text=self.s_label_Fdname + Fdpath[Fdpath.rfind("/", 0, Fdpath.rfind("/") - 1):])
            self.click_button_Refresh()
            E_file.cfg_ini("Fd_name_cfg", "w", valin=Fdpath)
        else:
            self.label_fd.config(text=self.s_label_Fdname + '你没有选择路径')

    # 刷新文件夹列表，文件名
    def click_button_Refresh(self):
        E_file.w_log("app", "刷新文件夹列表>>点击刷新文件列表")
        # lambda x=self.listbox_fd: x.delete()
        self.label_fd.config(text=self.s_label_Fdname + Fdpath[Fdpath.rfind("/", 0, Fdpath.rfind("/") - 1):])
        self.listbox_fd.delete(0, self.listbox_fd.size())  # (0, "end")
        items = self.pr.Listinfolder()
        for i in items:
            self.listbox_fd.insert("end", i)
        self.listbox_fd.see("end")  # (self.listbox_fd.size()"")
        self.listbox_fd.activate("end")
        self.reinput()
        self.n_activate = -1
        self.ListBox_var = ""
        self.set_label_filename()
        self.b_event_activate = False

    # 打开文件夹按键
    def click_button_openFd(self):
        E_file.w_log("app", "打开文件夹按键>>点击打开文件夹")
        self.listbox_fd.see(self.n_activate)
        self.listbox_fd.activate(self.n_activate)
        if self.ListBox_var == "" or not self.b_event_activate: messagebox.showwarning("提示", "未选择文件夹，将打开根目录")
        l_judge = self.pr.openFd(self.ListBox_var)
        if l_judge[0] == False: messagebox.showerror("错误", l_judge[1])
        print("打开文件夹")

    # 打开表格按键
    def click_button_open(self):
        E_file.w_log("app", "打开表格按键>>点击打开表格")
        l_judge = self.pr.openEXCEL(self.s_label_filename)
        if not l_judge[0]: messagebox.showerror("错误", l_judge[1])
        print("打开表格")

    # 重命名文件夹按键
    def click_button_reFd(self, from_ok=False):
        if not from_ok:
            E_file.w_log("app", "重命名文件夹按键>>点击重命名文件夹")
            self.n_sign = 0
            self.set_button_OKandCancel()
            self.button_formatFd.grid(row=6, column=4)
            self.n_sign = 1
        else:
            E_file.w_log("app", "重命名文件夹按键>>点击Ok")
            a = self.formatting_id(self.entry_id.get())
            b = self.formatting_serialnumber(self.entry_serialnumber.get())
            c = self.entry_problem.get()
            if c == "": messagebox.showerror("内容错误", "没有输入问题名")
            d = self.formatting_date(self.entry_date.get(), newFd=True)
            if b != "" and c != "" and d != "":
                self.pr.reFd(self.ListBox_var, a + "_" + b + " " + c + " " + d)
                return True
            else:
                return False

    # 新建问题/文件夹
    def click_button_newFd(self, from_ok=False):
        if not from_ok:
            E_file.w_log("app", "新建文件夹按键>>点击新建文件")
            self.n_sign = 0
            self.set_button_OKandCancel()
            self.reinput()
            namein = self.l_entry
            for i in range(0, len(namein)): namein[i] = ''
            namein[0] = self.formatting_id(str(self.pr.n_max_id + 1))
            namein[3] = self.formatting_id(self.auto_date())
            self.show_entry(namein)
            self.n_sign = 2
        else:
            E_file.w_log("app", "新建文件夹按键>>点击Ok")
            a = self.formatting_id(self.entry_id.get())
            b = self.formatting_serialnumber(self.entry_serialnumber.get())
            c = self.entry_problem.get()
            if c == "": messagebox.showerror("内容错误", "没有输入问题名")
            d = self.formatting_date(self.entry_date.get(), newFd=True)
            Fdname = a + "_" + b + " " + c + " " + d
            val_temp = self.getentry()
            check_temp = self.pr.set_Xlsx("open")
            check_temp1 = self.pr.write_el(val_temp, True)
            if check_temp[0] and check_temp1[0]:
                check_temp3 = self.pr.set_Xlsx("save")
                if check_temp3[0]:
                    if b != "" and c != "" and d != "":
                        self.pr.newFd(Fdname)
                        return True
                    else:
                        return False
                else:
                    messagebox.showerror("表格保存错误", "请检查：\n" + check_temp3[1])
            elif not check_temp[0]:
                messagebox.showerror("表格打开错误", "请检查：\n" + check_temp[1])
                return False
            elif not check_temp1[0]:
                messagebox.showerror("表格写入错误", "请检查：\n" + check_temp1[1])
                return False

    # 编辑表格内容
    def click_button_editor(self, from_ok=False):
        if not from_ok:
            E_file.w_log("app", "编辑按键>>点击编辑")
            self.n_sign = 0
            self.set_button_OKandCancel()
            self.n_sign = 3
        else:
            E_file.w_log("app", "编辑按键>>点击Ok")
            val_temp = self.getentry()
            if val_temp[0] == '':
                messagebox.showerror("名称错误", "未选择文件夹")
                return False
            check_temp = self.pr.set_Xlsx("open")
            check_temp1 = self.pr.write_el(val_temp, False)
            if check_temp[0] and check_temp1[0]:
                check_temp3 = self.pr.set_Xlsx("save")
                if check_temp3[0]:
                    return True
                else:
                    messagebox.showerror("表格保存错误", "请检查：\n" + check_temp3[1])
            elif not check_temp[0]:
                messagebox.showerror("表格打开错误", "请检查：\n" + check_temp[1])
                return False
            elif not check_temp1[0]:
                messagebox.showerror("表格写入错误", "请检查：\n" + check_temp1[1])
                return False

    # 获取输入内容
    def getentry(self):
        list_temp = []
        for s in self.list_entry: list_temp.append(eval('self.{}.get()'.format(s)))
        val_temp = [list_temp[0], self.pr.newpath + "\\" + self.ListBox_var, list_temp[1], "", list_temp[2],
                    list_temp[3], self.text_content.get(1.0, "end")[:-1], list_temp[4], list_temp[5], list_temp[7],
                    list_temp[6], list_temp[8]]
        return val_temp

    # 清屏
    def reinput(self):
        self.entry_id.configure(state="normal")
        self.entry_id.delete(0, "end")
        self.entry_id.configure(state="disabled")
        self.entry_serialnumber.delete(0, "end")
        self.entry_problem.delete(0, "end")
        self.entry_raiseAquestion.delete(0, "end")
        self.entry_ProblemHandling.delete(0, "end")
        self.entry_date.delete(0, "end")
        self.entry_state.delete(0, "end")
        self.entry_frequency.delete(0, "end")
        self.entry_remark.delete(0, "end")
        self.text_content.delete(1.0, "end")

    # 显示输入框内容
    def show_entry(self, lists):
        self.entry_id.configure(state="normal")
        self.entry_id.insert(0, lists[0])
        self.entry_id.configure(state="disabled")
        self.entry_serialnumber.insert(0, lists[1])
        self.entry_problem.insert(0, lists[2])
        self.entry_date.insert(0, lists[3])
        self.entry_raiseAquestion.insert(0, lists[4])
        self.entry_ProblemHandling.insert(0, lists[5])
        self.entry_state.insert(0, lists[7])
        self.entry_frequency.insert(0, lists[6])
        self.entry_remark.insert(0, lists[8])

    # 显示文本框内容
    def show_text(self, i):
        """显示文本框内容:参数 i = 字符串"""
        self.text_content.insert(1.0, i)
        self.text_content.see("end")

    # 调用顶级窗口
    def run_top(self):
        if self.b_top_living: self.top.destroy()
        self.set_init_toplevel()
        self.top.deiconify()
        self.s_label_filename = ""
        items = self.pr.get_file_list()
        self.listbox_f.delete(0, "end")
        for i in items:
            self.listbox_f.insert("end", i)

    # 设置顶级窗口
    def set_init_toplevel(self):
        # 主窗口属性
        self.b_top_living = False
        self.top = Toplevel()
        self.top.title("选择表格文件")
        self.top.geometry("288x140+600+240")
        self.top.resizable(0, 0)
        # 元素标签
        self.label_filename_from_top = Label(self.top, text="未选择文件", anchor="w", width=35)
        self.label_filename_from_top.grid(row=1, column=0, sticky="w", columnspan=3)
        # 元素列表
        self.listbox_f = Listbox(self.top, height=5, width=40)
        self.listbox_f.grid(row=0, column=0, columnspan=2, pady=3)
        self.listbox_f.bind("<ButtonRelease-1>", self.Event_listbox_var_form_top)
        # 元素按键
        self.button_Ok_from_top = Button(self.top, text="确认", width=10, command=self.click_button_Ok_from_top, bd=2)
        self.button_Ok_from_top.grid(row=1, column=1, sticky="e")
        # 默认关闭显示
        self.top.withdraw()
        self.b_top_living = True

    # top 列表框事件
    def Event_listbox_var_form_top(self, event):
        n_activate = self.listbox_f.curselection()
        if n_activate != ():
            self.s_label_filename = self.listbox_f.get(n_activate)
            # self.E_file.w_log()
            self.listbox_f.see(n_activate)
            self.label_filename_from_top.config(text=self.s_label_filename)

    # top 按键事件
    def click_button_Ok_from_top(self):
        n_activate = self.listbox_f.curselection()
        if self.s_label_filename != "":
            self.listbox_f.see(n_activate)
            self.top.withdraw()  # 关闭窗口（不销毁）
            self.top.destroy()  # 销毁窗口
            self.pr.s_xlsx_name = self.s_label_filename
            self.label_filename.config(text=self.s_label_filename)
            E_file.cfg_ini("Xlsx_name_cfg", "w", valin=self.s_label_filename, Fdnamein=Fdpath)
        else:
            messagebox.showerror("错误", "你未选择文件")
            self.top.deiconify()


class E_file:
    log_name: str = ""
    cfg_name: str = ""

    def __init__(self):
        self.log_name = self.new_file("journal", "ProblemManagement", "log")[1]

    @staticmethod  # 新建log文件
    def new_file(Fdnamein="/", namein="new", suffixin="txt", b_time=True):
        decode_encode = E_file.decode_encode
        if Fdnamein == "/":
            newpath = Workpath + "\\"
        elif Fdnamein.find("/") == -1:
            newpath = Workpath + "\\" + Fdnamein + "\\"
        else:
            newpath = Fdnamein + "\\"
        if not (os.path.exists(newpath)): os.mkdir(newpath)
        if b_time:
            s_time = " " + str(datetime.datetime.now())[:10]
        else:
            s_time = ""
        file_name = newpath + namein + s_time + "." + suffixin
        if not (os.path.exists(file_name)):
            open(file_name, mode="x", encoding="utf-8")  # x 创建 w 写入（清空文件，无文件将创建）
            return True, file_name
        else:
            return False, file_name

    @staticmethod  # 写入log方法
    def w_log(tepy, var, sep="更改>>", count=30):
        str_conduct = E_file.str_conduct
        decode_encode = E_file.decode_encode
        s_time = datetime.datetime.now()
        l_name = E_file.new_file("journal", "ProblemManagement", "log")
        E_file.log_name = l_name[1]
        if l_name[0]:
            f = open(E_file.log_name, mode="w", encoding="utf-8")
            f.writelines([f"日志名：{E_file.log_name}\n", f"创建日期：{str(s_time)[:19]}\n\n",
                          f"{'日期':^26} {decode_encode('类型', 6, '^')}   内容\n"])
            f.close()
        f = open(E_file.log_name, mode="a", encoding="utf-8")
        # print('This will be written to somedir/spamspam.txt', file=f)
        if var.find(sep) != -1:
            var1, var2 = var[:var.find(sep)], var[var.find(sep) + len(sep):]
            str_conduct(f, s_time, tepy, var1, var2, sep, count)
        else:
            b_one = False
            varin_temp = var + "\nend"
            while varin_temp.find("\n") != -1:
                var_temp = varin_temp[:varin_temp.find('\n')]
                if not b_one:
                    f.write(
                        f"{str(s_time):28} {decode_encode(tepy, 6, '^')}   {var_temp}\n")
                    b_one = True
                else:
                    f.write(44 * " " + f"{var_temp}\n")
                varin_temp = varin_temp[varin_temp.find("\n") + 1:]
        # f.write(f"{str(s_time):^20} {tepy:^10} {var:<60}\n")
        f.close()

    @staticmethod
    def decode_encode(strin, numin, align="^"):
        """align = ^:居中 <:靠左 >:靠右"""
        numin = numin * 2
        n_max, n_head, n_end, n_mend = 0, 0, 0, 0
        for s in strin:
            if len(s.encode()) != 3:
                n_max = n_max + 1
            else:
                n_max = n_max + 2
        n_mend = numin - n_max
        if align == "^":
            n_head = n_mend // 2
            if n_mend % 2 == 0:
                n_end = n_head
            else:
                n_end = n_mend // 2 + 1
        elif align == "<":
            n_end = n_mend
        elif align == ">":
            n_head = n_mend
        else:
            n_head, n_end = 0, 0
        strout: str = n_head * " " + strin + n_end * " "
        return strout

    '''        
        b_strin = strin.encode()
        if len(b_strin) % 2 != 0: b_strin = b_strin + b" "
        n_max = (numin - len(b_strin)) // 3  # +(numin-len(b_strin))%3
        if n_max % 2 == 0:
            n_max = n_max // 2
        else:
            n_max = n_max // 2 + 1
        if align == "^":
            n_head = n_max // 2*4
            if n_max % 2 == 0:
                n_end = n_head
            else:
                n_end = n_max // 2*4 + 1
        elif align == "<":
            n_end = n_max
        elif align == ">":
            n_head = n_max
        else:
            n_head, n_end = 0, 0
        # b_strout = n_head * b" " + b_strin + n_end * b" "
        # return b_strout.decode()
        strout:str = n_head * " " + strin + n_end * " "
        return strout
    '''

    @staticmethod
    def str_conduct(f, s_time, tepy, var, var1, sep="", count=30):
        b_one = False
        decode_encode = E_file.decode_encode
        if sep == "":
            varin_temp = var + "\nend"
            while varin_temp.find("\n") != -1:
                var_temp = varin_temp[:varin_temp.find('\n')]
                if len(var_temp) > count:
                    while len(var_temp) > count:
                        var_temp, var_temp_end = var_temp[:count], var_temp[count:]
                        if not b_one:
                            f.write(
                                f"{str(s_time):28} {decode_encode(tepy, 6, '^')}   {decode_encode(var_temp, count, '<')}\n")
                            b_one = True
                        else:
                            f.write(44 * " " + f"{decode_encode(var_temp, count, '<')}\n")
                        var_temp = var_temp_end
                    f.write(44 * " " + f"{decode_encode(var_temp, count, '<')}\n")
                else:
                    if not b_one:
                        f.write(
                            f"{str(s_time):28} {decode_encode(tepy, 6, '^')}   {decode_encode(var_temp, count, '<')}\n")
                        b_one = True
                    else:
                        f.write(44 * " " + f"{decode_encode(var_temp, count, '<')}\n")
                varin_temp = varin_temp[varin_temp.find("\n") + 1:]
            return
        varin_temp, varin_temp1, var_temp, var_temp1 = var + "\nend", var1 + "\nend", "", ""
        while varin_temp.find("\n") != -1 or varin_temp1.find("\n") != -1:
            if varin_temp.find("\n") != -1:
                var_temp = varin_temp[:varin_temp.find('\n')]
            else:
                varin_temp = ""
            if varin_temp1.find("\n") != -1:
                var_temp1 = varin_temp1[:varin_temp1.find('\n')]
            else:
                varin_temp1 = ""
            if len(var_temp) > count or len(var_temp1) > count:
                while len(var_temp) > count or len(var_temp1) > count:
                    if len(var_temp) > count:
                        var_temp, var_temp_end = var_temp[:count], var_temp[count:]
                    else:
                        var_temp, var_temp_end = var_temp, ""
                    if len(var_temp1) > count:
                        var_temp1, var_temp_end1 = var_temp1[:count], var_temp1[count:]
                    else:
                        var_temp1, var_temp_end1 = var_temp1, ""
                    if not b_one:
                        f.write(
                            f"{str(s_time):28} {decode_encode(tepy, 6, '^')}   {decode_encode(var_temp, count, '<')} {sep} {decode_encode(var_temp1, count, '<')}\n")  # {decode_encode(var_temp, 30, '<')} {sep} {decode_encode(var_temp1, 30, '<')}\n")
                        b_one = True
                    else:
                        f.write(
                            44 * " " + f"{decode_encode(var_temp, count, '<')} {sep} {decode_encode(var_temp1, count, '<')}\n")
                    var_temp, var_temp1 = var_temp_end, var_temp_end1
                f.write(
                    44 * " " + f"{decode_encode(var_temp, count, '<')} {sep} {decode_encode(var_temp1, count, '<')}\n")
            else:
                if not b_one:
                    f.write(
                        f"{str(s_time):28} {decode_encode(tepy, 6, '^')}   {decode_encode(var_temp, count, '<')} {sep} {decode_encode(var_temp1, count, '<')}\n")
                    b_one = True
                else:
                    f.write(
                        44 * " " + f"{decode_encode(var_temp, count, '<')} {sep} {decode_encode(var_temp1, count, '<')}\n")
            if varin_temp.find("\n") != -1:
                varin_temp = varin_temp[varin_temp.find("\n") + 1:]
            else:
                varin_temp = varin_temp
            if varin_temp1.find("\n") != -1:
                varin_temp1 = varin_temp1[varin_temp1.find("\n") + 1:]
            else:
                varin_temp1 = varin_temp1

    @staticmethod
    def cfg_ini(par, mode="r", *, valin="", Fdnamein="/", namein="ProblemManagementCfg", suffixin="ini", b_time=False):
        str_conduct = E_file.str_conduct
        decode_encode = E_file.decode_encode
        s_time = datetime.datetime.now()
        l_name = E_file.new_file(Fdnamein, namein, suffixin, b_time)
        E_file.cfg_name = l_name[1]
        be, val, val_num = False, "", 0
        match mode:
            case "r":
                with open(E_file.cfg_name, mode="r", encoding="utf-8") as rf:
                    for i in rf.readlines():
                        if i[i.find("[") + 1:i.find("]")] == par:
                            val = i[i.find(":") + 1:i.find("\n")]
                            if val != "": be = True
            case "w":
                b_val_be = False
                with open(E_file.cfg_name, mode="r", encoding="utf-8") as rf:
                    data_temp = rf.readlines()
                    for i in range(0, len(data_temp)):
                        if data_temp[i][data_temp[i].find("[") + 1:data_temp[i].find("]")] == par:
                            b_val_be, s, val = True, data_temp[i], str(i)
                            data_temp[i] = "{}:{}\n".format(s[:s.find(":")], valin)
                with open(E_file.cfg_name, mode="w", encoding="utf-8") as wf:
                    wf.writelines(data_temp)
                if not b_val_be:
                    with open(E_file.cfg_name, mode="a", encoding="utf-8") as af:
                        af.write(f"[{par}]:{valin}\n\n")
                be = True
        return be, val


class Fd_editor():
    global Workpath
    global Fdpath

    # num 变量
    n_max_id = 0  # 当前最大编号

    # str 变量
    s_xlsx_name = ""  # 表格文件名

    def __init__(self):
        # self.log = log_app()
        self.set_Fdpath()

    # 设置工作路径
    def set_Fdpath(self):
        if Fdpath == "":
            self.newpath = Workpath
        else:
            self.newpath = Fdpath

    # 设置表格对象
    def set_Xlsx(self, operate="open"):
        try:
            # if operate != "open" or "close" or "save":raise NameError()
            if operate == "open":
                self.set_Fdpath()
                self.ex = openpyxl.load_workbook(self.newpath + "\\" + self.s_xlsx_name)  # , data_only=True)
                self.sh = self.ex.active
                if str(self.sh.cell(3, 1).value)[:1] == "=":
                    self.ex = openpyxl.load_workbook(self.newpath + "\\" + self.s_xlsx_name, data_only=True)
                    self.sh = self.ex.active
                E_file.w_log("设置表格", "表格打开成功>" + self.s_xlsx_name)
            elif operate == "close":
                self.ex.close()
                E_file.w_log("设置表格", "表格关闭成功>" + self.s_xlsx_name)
            elif operate == "save":
                self.ex.save(self.newpath + "\\" + self.s_xlsx_name)
                E_file.w_log("设置表格", "表格保存成功>" + self.s_xlsx_name)
            else:
                raise NameError()
            return True, 'Ok'
        except FileNotFoundError as error:
            E_file.w_log("错误", "设置表格>>文件未找到: " + str(error))
            return False, str(error)
        except NameError as error:
            E_file.w_log("错误", "设置表格>>参数错误: ")
            return False, "参数错误"
        except PermissionError as error:
            E_file.w_log("错误", "设置表格>>无文件处理权限: " + str(error))
            return False, '文件保存失败:' + str(error) + '\n请关闭已打开文件后重试'

    '''
        except openpyxl.utils.exceptions.InvalidFileException as error:
            E_file.w_log("错误", "设置表格>>utils: " + str(error))
            return False, str(error)
    '''

    # 重命名文件夹
    def reFd(self, name, new_name):
        try:
            self.set_Fdpath()
            os.rename(self.newpath + "\\" + name, self.newpath + "\\" + new_name)
            E_file.w_log("重命名文件夹", name + " 命名为: " + new_name)
            return True, 'Ok'
        except OSError as error:
            E_file.w_log("错误", "重命名文件夹>>系统: " + str(error))
            return False, str(error)

    # 新建文件夹
    def newFd(self, name):
        try:
            self.set_Fdpath()
            os.mkdir(self.newpath + '\\' + name)
            E_file.w_log("新建文件夹", "在路径：" + self.newpath + " 下创建文件夹: " + name)
            return True, "Ok"
        except OSError as error:
            E_file.w_log("错误", "创建文件夹>>系统: " + str(error))
            return False, str(error)

    # 获取文件夹列表
    def Listinfolder(self):
        self.set_Fdpath()
        getlist = next(os.walk(self.newpath))[1]
        outlist = self.fromdisposelist(getlist)
        E_file.w_log("文件目录", "获取" + self.newpath + "下文件夹列表")
        return outlist

    # 处理特定的文件夹进入列表
    def fromdisposelist(self, newlist):
        n_list = []
        self.n_max_id = 0
        newlist.sort()
        for h in newlist:
            x = h.find('_')
            if x != 0:
                if x != -1:
                    s = h[:x]
                    try:
                        s = int(s)
                        if s > self.n_max_id: self.n_max_id = s
                        n_list.insert(s, h)
                    except:
                        pass
        return n_list

    # 打开文件夹
    def openFd(self, fdname):
        try:
            self.set_Fdpath()
            os.startfile(self.newpath + "\\" + fdname)
            # os.system('start explorer ' + self.newpath + "\\" + fdname)
            E_file.w_log("文件夹操作", "打开" + self.newpath + "下文件夹: " + fdname)
            return True, "Ok"
        except OSError as error:
            E_file.w_log("错误", "打开文件夹>>系统: " + str(error))
            return False, str(error)

    # 打开表格
    def openEXCEL(self, name):
        try:
            self.set_Fdpath()
            if name == "": raise NameError()
            os.startfile(self.newpath + "\\" + name)
            # os.system('start EXCEL.exe ' + '"' + self.newpath + "\\" + name + '"')
            E_file.w_log("打开表格", "打开" + self.newpath + "下表格: " + name)
            return True, "Ok"
        except OSError as error:
            E_file.w_log("错误", "打开表格>>系统: " + str(error))
            return False, str(error)
        except NameError:
            E_file.w_log("错误", "打开表格>>系统: 没有指定表格名")
            return False, "没有选择要打开的文件"

    # 获取指定类型文件列表
    def get_file_list(self, dis_name="xlsx"):
        self.set_Fdpath()
        # self.E_file.w_log()
        getlist = next(os.walk(self.newpath))[2]
        outlist = []
        nm = 0
        for i in getlist:
            if i.find("~$") != -1: continue
            pos = i.rfind(".")
            if 0 < pos < len(i) - 1:
                if i[pos + 1:] == dis_name:
                    outlist.append(i)
                    nm += 1
        E_file.w_log("文件目录", "获取" + self.newpath + "下: ." + dis_name + "文件列表")
        if outlist:
            return outlist
        else:
            return ["无有效文件"]

    # 读取表格编号行内容
    def read_el(self, id):
        out_val = []
        for r in range(3, self.sh.max_row + 1):
            try:
                id_temp = int(self.sh.cell(r, 1).value)
            except:
                continue
            if id_temp == int(id):
                for c in range(1, self.sh.max_column + 1):
                    if self.sh.cell(r, c).value is None:
                        out_val.append("")
                    else:
                        out_val.append(self.sh.cell(r, c).value)
        if not out_val:
            E_file.w_log("错误", "读取" + self.s_xlsx_name + "文件出错: " + str(id) + "行不存在")
            return [x for x in 9 * " "]
        else:
            E_file.w_log("读文件", "读取" + self.s_xlsx_name + "文件：" + str(id) + "行内容")
            return out_val

    # 写入表格编号行内容
    def write_el(self, val, new=False):
        if new:
            id_row = int(val[0]) + 2
            if self.sh.cell(id_row, 1).value == " " or self.sh.cell(id_row, 1).value is None:
                E_file.w_log("写文件", "写入" + self.s_xlsx_name + "文件: " + str(id_row) + "行:")
                for c in range(1, self.sh.max_column + 1):
                    if c <= len(val) and val[c - 1] != "":
                        if c == 1:
                            self.sh.cell(id_row, c, int(val[c - 1]))
                        elif c == 2:
                            self.sh.cell(id_row, c).hyperlink = val[c - 1]
                            self.sh.cell(id_row, c, val[c - 1][val[c - 1].rfind("\\") + 1:])
                        else:
                            self.sh.cell(id_row, c, val[c - 1])
                        E_file.w_log("写文件", str(c) + "列: " + val[c - 1])
                return True, "Ok"
            else:
                E_file.w_log("错误", "写入" + self.s_xlsx_name + "文件出错: " + str(int(val[0]) + 2) + "行不为空")
                return False, "所选编号在表格位置中不为空，表格编号：" + str(self.sh.cell(int(val[0]) + 2, 1).value)
        else:
            E_file.w_log("写文件", "写入" + self.s_xlsx_name + "文件: " + str(int(val[0]) + 2) + "行:")
            for r in range(3, self.sh.max_row + 1):
                if self.sh.cell(r, 1).value == int(val[0]):
                    for c in range(1, self.sh.max_column + 1):
                        if c <= len(val) and val[c - 1] != "":
                            var_former = self.sh.cell(r, c).value
                            if c == 1:
                                self.sh.cell(r, c, int(val[c - 1]))
                            elif c == 2:
                                self.sh.cell(r, c).hyperlink = val[c - 1]
                                self.sh.cell(r, c, val[c - 1][val[c - 1].rfind("\\") + 1:])
                            else:
                                self.sh.cell(r, c, val[c - 1])
                            if str(var_former) != str(val[c - 1]):
                                E_file.w_log("写文件", str(var_former) + " 更改>> " + str(val[c - 1]), " 更改>> ")
            return True, "Ok"


'''
    # 删除文件夹
    def delefolder(name):
        os.removedirs(Workpath + '\\' + name)
        print('删除成功')

class app1(app):
    def click_button_open(self):
        self.pr.tss()
        print("打开表格")
'''

a = app().run()
